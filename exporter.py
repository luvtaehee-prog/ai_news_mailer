"""데이터 내보내기 모듈 (4번 파트).

역할: clean 뉴스와 AI 요약 결과를 합쳐 CSV / JSONL / Excel 파일로 내보낸다.

- clean 저장소(news_clean.jsonl)와 요약 저장소(news_summary.jsonl)를
  뉴스 id 기준으로 합쳐 한 줄에 '기사 + 요약'이 모두 담기게 만든다.
- 요약이 없는 기사도 status="unsummarized" 로 남겨서 진행 상황을 알 수 있게 한다.

단독 실행:
    python exporter.py --format csv
"""

import os
import csv
import json
from datetime import datetime

from log_setup import get_logger

log = get_logger("report")

# 내보내기 컬럼 순서 (엑셀에서 바로 읽기 좋은 순서로 배치)
COLUMNS = [
    "id", "published_date", "category", "source", "press", "title",
    "summary", "keywords", "status", "content_length", "truncated",
    "url", "collected_at",
]

SUPPORTED_FORMATS = ["csv", "jsonl", "excel"]


def build_rows(clean: list, summaries: list) -> list:
    """clean 뉴스와 요약을 id 기준으로 합쳐 표 형태(행 리스트)로 만든다."""
    summary_by_id = {s.get("id"): s for s in summaries if s.get("id")}

    rows = []
    for rec in clean:
        rid = rec.get("id")
        summary = summary_by_id.get(rid, {})
        keywords = summary.get("keywords") or []
        rows.append({
            "id": rid,
            "published_date": rec.get("published_date", ""),
            "category": rec.get("category", ""),
            "source": rec.get("source", ""),
            "press": rec.get("press", ""),
            "title": rec.get("title", ""),
            "summary": summary.get("summary", ""),
            # 리스트를 그대로 넣으면 CSV/엑셀에서 읽기 어려워 쉼표로 이어 붙인다
            "keywords": ", ".join(keywords),
            "status": "summarized" if summary else "unsummarized",
            "content_length": rec.get("content_length",
                                      len(rec.get("content") or "")),
            "truncated": rec.get("truncated", False),
            "url": rec.get("url", ""),
            "collected_at": rec.get("collected_at", ""),
        })

    # 최신 기사가 위로 오도록 발행일 내림차순 정렬
    rows.sort(key=lambda r: r["published_date"], reverse=True)
    return rows


def filter_rows(rows: list, status: str = None, category: str = None,
                date_from: str = None, date_to: str = None,
                limit: int = None) -> list:
    """내보내기 대상 행을 조건에 맞게 걸러낸다."""
    out = []
    for r in rows:
        if status and status != "all" and r["status"] != status:
            continue
        if category and r["category"] != category:
            continue
        if date_from and r["published_date"] < date_from:
            continue
        if date_to and r["published_date"] > date_to:
            continue
        out.append(r)
    if limit:
        out = out[:limit]
    return out


def _out_path(out_dir: str, fmt: str, filename: str = None) -> str:
    """저장 경로를 만든다 (파일명 미지정 시 날짜시각을 붙인다)."""
    os.makedirs(out_dir, exist_ok=True)
    ext = {"csv": "csv", "jsonl": "jsonl", "excel": "xlsx"}[fmt]
    # 초 단위까지 붙여 같은 분에 두 번 실행해도 이전 파일을 덮어쓰지 않게 한다
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(out_dir, filename or f"news_export_{stamp}.{ext}")


def export_csv(rows: list, path: str) -> str:
    """CSV 로 저장한다.

    인코딩은 utf-8-sig(BOM 포함)를 쓴다. 그래야 윈도우 엑셀에서
    파일을 열었을 때 한글이 깨지지 않는다.
    """
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS)
        writer.writeheader()
        writer.writerows(rows)
    log.info("CSV 저장: %s (%d행)", path, len(rows))
    return path


def export_jsonl(rows: list, path: str) -> str:
    """JSONL 로 저장한다 (한 줄에 뉴스 하나)."""
    with open(path, "w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")
    log.info("JSONL 저장: %s (%d행)", path, len(rows))
    return path


def export_excel(rows: list, path: str, stats: dict = None) -> str:
    """Excel(.xlsx) 로 저장한다. stats 를 주면 '요약 통계' 시트를 추가한다."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.error("openpyxl 이 설치되어 있지 않습니다. "
                  "pip install openpyxl 후 다시 실행하세요.")
        raise

    wb = Workbook()
    ws = wb.active
    ws.title = "뉴스"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="3B6FD4")

    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        ws.append([row.get(col, "") for col in COLUMNS])

    # 컬럼별 보기 좋은 너비 (긴 텍스트 컬럼은 넓게)
    widths = {"title": 50, "summary": 70, "keywords": 30, "url": 45,
              "collected_at": 22, "published_date": 14}
    for idx, col in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(col, 14)

    ws.freeze_panes = "A2"          # 헤더 고정
    ws.auto_filter.ref = ws.dimensions  # 열 필터 버튼 추가

    if stats:
        ws2 = wb.create_sheet("요약 통계")
        ws2.append(["지표", "값"])
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
        for k, v in stats.get("quality", {}).items():
            ws2.append([k, str(v)])
        ws2.append([])
        ws2.append(["카테고리", "건수"])
        for k, v in stats.get("category_counts", {}).items():
            ws2.append([k, v])
        ws2.append([])
        ws2.append(["키워드", "등장 기사 수"])
        for kw, c in stats.get("top_keywords", []):
            ws2.append([kw, c])
        ws2.column_dimensions["A"].width = 26
        ws2.column_dimensions["B"].width = 34

    wb.save(path)
    log.info("Excel 저장: %s (%d행)", path, len(rows))
    return path


def export(rows: list, fmt: str, out_dir: str = "output/exports",
           filename: str = None, stats: dict = None) -> str:
    """포맷에 맞는 내보내기 함수를 호출한다."""
    if fmt not in SUPPORTED_FORMATS:
        raise ValueError(f"지원하지 않는 포맷: {fmt} "
                         f"(사용 가능: {', '.join(SUPPORTED_FORMATS)})")

    path = _out_path(out_dir, fmt, filename)
    if fmt == "csv":
        return export_csv(rows, path)
    if fmt == "jsonl":
        return export_jsonl(rows, path)
    return export_excel(rows, path, stats)


if __name__ == "__main__":
    # 단독 실행: clean + 요약을 합쳐 CSV 로 내보낸다.
    import argparse

    from reporter import load_all

    parser = argparse.ArgumentParser(description="뉴스 데이터 내보내기")
    parser.add_argument("--format", default="csv", choices=SUPPORTED_FORMATS)
    parser.add_argument("--status", default="all",
                        choices=["all", "summarized", "unsummarized"])
    args = parser.parse_args()

    _data = load_all()
    _rows = filter_rows(build_rows(_data["clean"], _data["summaries"]),
                        status=args.status)
    print(export(_rows, args.format))
